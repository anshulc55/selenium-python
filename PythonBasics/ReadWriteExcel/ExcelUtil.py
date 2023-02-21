# Write the functions to perform some common Operations on Excel Data

import openpyxl
import os


class ReadWriteExcel:
    wb = None
    file_path = None
    file_dir = os.getcwd() + "/FileData/"

    def __init__(self, filename):
        global wb
        global file_dir
        global file_path
        file_path = self.file_dir + filename
        wb = openpyxl.load_workbook(file_path)
        self.file_path = file_path

    # Read Value by Column Index
    def readby_columnIndex(self, sheetname, rowNum, colNum):
        sheet = wb[sheetname]
        return sheet.cell(rowNum, colNum).value

    # read Value by Column Name
    def readby_columnName(self, sheetname, rowNum, colName):
        sheet = wb[sheetname]
        colIndex = 1
        while sheet.cell(row=1, column=colIndex).value != "":
            if sheet.cell(row=1, column=colIndex).value == colName:
                break
            colIndex = colIndex + 1
        return sheet.cell(rowNum, colIndex).value

    # To get the total Number of Rows
    def get_totalRows(self, sheetname):
        sheet = wb[sheetname]
        return sheet.max_row

    # To get the Total Number of Column
    def get_totalColumn(self, sheetname):
        sheet = wb[sheetname]
        return sheet.max_column

    # To write by Column Index
    def writeby_columIndex(self, sheetname, rowNum, colNum, value):
        sheet = wb[sheetname]
        sheet.cell(row=rowNum, column=colNum).value = value
        wb.save(file_path)










