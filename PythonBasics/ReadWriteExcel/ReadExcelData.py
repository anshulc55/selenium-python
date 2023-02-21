import openpyxl
import os

file_path = os.getcwd()+"/FileData/StudentData.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# Read Data from cell
'''
print(sheet["A5"].value)
print(sheet["C3"].value)
print(sheet.cell(row=6, column=4).value)
'''

# Read multiple cells
'''
cells = sheet["A1":"C5"]

for a1, a2, a3 in cells:
    print("{0:6} {1:6} {2:6}".format(a1.value, a2.value, a3.value))
'''

# Determine total number of rows
#print(sheet.max_row)

# Determine total number of columns
#print(sheet.max_column)

# Print all columns name
'''
for col in range(1, sheet.max_column+1):
    cell = sheet.cell(row=1, column=col)
    print(cell.value)
'''

# Print a particular row value
'''
for col in range(1, sheet.max_column+1):
    cell = sheet.cell(row=6, column=col)
    print(cell.value, end=" ")
'''

# Iterate by rows
'''
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=7, max_col=4):
    for cell in row:
        print(cell.value, end=" ")
    print()
'''

# Iterate by Column
for row in sheet.iter_cols(min_row=1, min_col=1, max_row=6, max_col=4):
    for cell in row:
        print(cell.value, end=" ")
    print()
